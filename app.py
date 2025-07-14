import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
from io import BytesIO
from geopy.distance import geodesic
import openpyxl
from typing import Dict, List, Optional, Union

# Initialize session state at the beginning
def initialize_session_state():
    """Initialize all session state variables"""
    if 'boq_form_values' not in st.session_state:
        st.session_state.boq_form_values = {
            'lop_name': "",
            'sumber': "ODC",
            'kabel_12': 0.0,
            'kabel_24': 0.0,
            'odp_8': 0,
            'odp_16': 0,
            'tiang_new': 0,
            'tiang_existing': 0,
            'tikungan': 0,
            'izin': "",
            'closure': 0,
            'otb_12': 0,
            'uploaded_file': None,
            'kml_file': None
        }

    if 'boq_state' not in st.session_state:
        st.session_state.boq_state = {
            'ready': False,
            'excel_data': None,
            'project_name': "",
            'updated_items': [],
            'summary': {},
            'active_tab': "manual"
        }

def reset_boq_application():
    """Reset the application to its initial state"""
    st.session_state.boq_form_values = {
        'lop_name': "",
        'sumber': "ODC",
        'kabel_12': 0.0,
        'kabel_24': 0.0,
        'odp_8': 0,
        'odp_16': 0,
        'otb_12': 0,
        'tiang_new': 0,
        'tiang_existing': 0,
        'tikungan': 0,
        'izin': "",
        'closure': 0,
        'uploaded_file': None,
        'kml_file': None
    }
    st.session_state.boq_state = {
        'ready': False,
        'excel_data': None,
        'project_name': "",
        'updated_items': [],
        'summary': {},
        'active_tab': "manual"
    }

def validate_kml_file(kml_data: bytes) -> bool:
    """Validate the KML file structure"""
    try:
        ET.fromstring(kml_data)
        return True
    except ET.ParseError:
        return False

def parse_kml_file(kml_file) -> Optional[Dict[str, Union[int, float]]]:
    """Parse KML file and extract relevant values"""
    try:
        kml_data = kml_file.read()
        if not kml_data:
            st.error("KML file is empty")
            return None
            
        if not validate_kml_file(kml_data):
            st.error("Invalid KML file format")
            return None
            
        root = ET.fromstring(kml_data)
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        
        values = {
            'tiang_new': 0,
            'tiang_existing': 0,
            'kabel_12': 0.0,
            'odp_8': 0,
            'odp_16': 0,
            'closure': 0,
            'otb_12': 0
        }
        
        for placemark in root.findall('.//kml:Placemark', ns):
            name_elem = placemark.find('kml:name', ns)
            desc_elem = placemark.find('kml:description', ns)
            
            name = name_elem.text.upper().strip() if name_elem is not None and name_elem.text else ""
            desc = desc_elem.text.strip() if desc_elem is not None and desc_elem.text else ""
            
            if placemark.find('.//kml:Point', ns) is not None:
                if any(keyword in name for keyword in ["TN", "TN7", "TIANG NEW"]):
                    values['tiang_new'] += 1
                elif any(keyword in name for keyword in ["TE", "TIANG EXISTING"]):
                    values['tiang_existing'] += 1
                elif "ODP" in name and any(keyword in name for keyword in ["NEW", "BARU"]):
                    if "8" in name or "ODP Solid-PB-8 AS" in desc:
                        values['odp_8'] += 1
                    elif "16" in name or "ODP Solid-PB-16 AS" in desc:
                        values['odp_16'] += 1
                elif "OTB" in name and any(keyword in name for keyword in ["NEW", "BARU"]):
                    values['otb_12'] += 1
                elif any(keyword in name for keyword in ["CL", "CLOSURE"]):
                    values['closure'] += 1
            
            elif placemark.find('.//kml:LineString', ns) is not None:
                if any(keyword in name for keyword in ["DIS NEW", "DISTRIBUSI", "AC-OF-SM-12"]):
                    coords_elem = placemark.find('.//kml:coordinates', ns)
                    if coords_elem is not None and coords_elem.text:
                        try:
                            coords = [
                                tuple(map(float, c.split(',')[:2]))
                                for c in coords_elem.text.split() 
                                if len(c.split(',')) >= 2
                            ]
                            if len(coords) > 1:
                                total_length = 0.0
                                for i in range(len(coords)-1):
                                    lon1, lat1 = coords[i]
                                    lon2, lat2 = coords[i+1]
                                    total_length += geodesic((lat1, lon1), (lat2, lon2)).meters
                                values['kabel_12'] += total_length
                        except ValueError:
                            continue
        
        return values
    
    except Exception as e:
        st.error(f"KML parsing failed: {str(e)}")
        return None

def calculate_puas(total_odp: int, tiang_new: int, tiang_existing: int, tikungan: int) -> int:
    """Calculate PU-AS volume"""
    return max(0, (total_odp * 2) - 1 + tiang_new + tiang_existing + tikungan)

def calculate_volumes(inputs: Dict) -> List[Dict]:
    """Calculate all required volumes based on inputs"""
    total_odp = inputs['odp_8'] + inputs['odp_16']
    
    # Calculate cable volumes with 2% buffer
    vol_kabel_12 = round(inputs['kabel_12'] * 1.02) if inputs['kabel_12'] > 0 else 0
    vol_kabel_24 = round(inputs['kabel_24'] * 1.02) if inputs['kabel_24'] > 0 else 0
    
    vol_puas = calculate_puas(total_odp, inputs['tiang_new'], inputs['tiang_existing'], inputs['tikungan'])
    
    # Source-specific calculations
    if inputs['sumber'] == "ODC":
        vol_os_sm_1_odc = total_odp * 2
        vol_os_sm_1_odp = 0
        vol_base_tray = 1 if vol_kabel_12 > 0 else 2 if vol_kabel_24 > 0 else 0
        vol_tc_02_odc = 1
        vol_dd_hdpe = 6
        vol_bc_tr = 3
    else:
        vol_os_sm_1_odc = 0
        vol_os_sm_1_odp = total_odp * 2
        vol_base_tray = 0
        vol_tc_02_odc = 0
        vol_dd_hdpe = 0
        vol_bc_tr = 0
    
    # Common calculations
    vol_os_sm_1 = vol_os_sm_1_odc + vol_os_sm_1_odp
    vol_pc_upc = ((total_odp - 1) // 4) + 1 if total_odp > 0 else 0
    vol_pc_apc = 18 if vol_pc_upc == 1 else vol_pc_upc * 2 if vol_pc_upc > 1 else 0
    vol_ps_1_4_odc = ((total_odp - 1) // 4) + 1 if total_odp > 0 else 0
    vol_ps_1_8_odp = 1 if inputs.get('otb_12', 0) > 0 else 0
    
    # Handle preliminary project cost
    izin_value = 0
    if inputs['izin']:
        try:
            izin_value = float(inputs['izin'].replace(',', '').replace('.', ''))
        except ValueError:
            st.warning("Invalid preliminary project value. Using 0 instead.")
    
    return [
        {"designator": "AC-OF-SM-12-SC_O_STOCK", "volume": vol_kabel_12},
        {"designator": "AC-OF-SM-24-SC_O_STOCK", "volume": vol_kabel_24},
        {"designator": "ODP Solid-PB-8 AS", "volume": inputs['odp_8']},
        {"designator": "ODP Solid-PB-16 AS", "volume": inputs['odp_16']},
        {"designator": "PU-S7.0-400NM", "volume": inputs['tiang_new']},
        {"designator": "PU-AS", "volume": vol_puas},
        {"designator": "OS-SM-1-ODC", "volume": vol_os_sm_1_odc},
        {"designator": "OS-SM-1-ODP", "volume": vol_os_sm_1_odp},
        {"designator": "OS-SM-1", "volume": vol_os_sm_1},
        {"designator": "PC-UPC-652-2", "volume": vol_pc_upc},
        {"designator": "PC-APC/UPC-652-A1", "volume": vol_pc_apc},
        {"designator": "PS-1-4-ODC", "volume": vol_ps_1_4_odc},
        {"designator": "TC-02-ODC", "volume": vol_tc_02_odc},
        {"designator": "DD-HDPE-40-1", "volume": vol_dd_hdpe},
        {"designator": "BC-TR-0.6", "volume": vol_bc_tr},
        {"designator": "Base Tray ODC", "volume": vol_base_tray},
        {"designator": "SC-OF-SM-24", "volume": inputs.get('closure', 0)},
        {"designator": "TC-SM-12", "volume": inputs.get('otb_12', 0)},
        {"designator": "PS-1-8-ODP", "volume": vol_ps_1_8_odp},
        {
            "designator": "Preliminary Project HRB/Kawasan Khusus",
            "volume": 1 if inputs['izin'] else 0,
            "izin_value": izin_value
        }
    ]

def process_boq_template(uploaded_file, inputs: Dict, lop_name: str) -> Optional[Dict]:
    """Process the BOQ template with calculated values"""
    try:
        with st.spinner("Processing BOQ template..."):
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active
            items = calculate_volumes(inputs)
            
            # Update worksheet with calculated values
            for row in range(9, 289):
                cell_value = str(ws[f'B{row}'].value or "").strip()
                
                for item in items:
                    if cell_value == item["designator"] and item["volume"] > 0:
                        ws[f'G{row}'] = item["volume"]
                        if "Preliminary" in cell_value and "izin_value" in item:
                            ws[f'F{row}'] = item["izin_value"]
            
            # Calculate totals
            material = jasa = 0.0
            for row in range(9, 289):
                try:
                    h_mat = float(ws[f'E{row}'].value or 0)
                    h_jasa = float(ws[f'F{row}'].value or 0)
                    vol = float(ws[f'G{row}'].value or 0)
                    material += h_mat * vol
                    jasa += h_jasa * vol
                except (ValueError, TypeError):
                    continue
            
            total = material + jasa
            total_odp = inputs['odp_8'] + inputs['odp_16']
            total_ports = (total_odp * 8) + (1 if inputs.get('otb_12', 0) > 0 else 0) * 8
            cpp = round(total / total_ports, 2) if total_ports > 0 else 0
            
            # Prepare output
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return {
                'excel_data': output,
                'summary': {
                    'material': material,
                    'jasa': jasa,
                    'total': total,
                    'cpp': cpp,
                    'total_odp': total_odp,
                    'total_ports': total_ports
                },
                'updated_items': [item for item in items if item['volume'] > 0]
            }
    
    except Exception as e:
        st.error(f"Error processing BOQ template: {str(e)}")
        return None

def display_kml_guide():
    """Display KML input guide in an expander"""
    with st.expander("📘 Panduan Input KML untuk BOQ Generator (⚠️readme first ya)"):
        st.markdown("""
**Berikut adalah aturan wajib untuk file KML yang akan diupload:**

### **1. Format File**
- File harus berformat **.kml** (tidak mendukung .kmz)
- Dapat dibuat menggunakan Google Earth, QGIS, atau software GIS lainnya

### **2. Penamaan Fitur Wajib**

#### **A. ODP (Optical Distribution Point)**
- Format: `ODP [Jenis] NEW/BARU`  
  Contoh:  
  ✅ `ODP 8 NEW`  
  ✅ `ODP 16 BARU`  
- Deskripsi wajib mencantumkan:  
  `ODP Solid-PB-8 AS` (untuk 8 port)  
  `ODP Solid-PB-16 AS` (untuk 16 port)  
- ❌ Tidak terbaca jika: Tidak ada keterangan NEW/BARU

#### **B. Tiang (Pole)**
- **Tiang Baru**:  
  ✅ `TN`, `TN7`, atau `TIANG NEW`  
  Contoh: `TN7-01 NEW`
- **Tiang Existing**:  
  ✅ `TE`, `TIANG EXISTING`, atau `TIANG EXIST`  
  Contoh: `TE-01`

#### **C. Kabel (Fiber)**
- Gunakan **LineString**  
- Nama Kabel New harus mengandung:  
  ✅ `DIS NEW`, `DISTRIBUSI`, `AC-OF-SM-12`, atau `BR`  
  Contoh: `DIS NEW JALAN A`
- Nama Kabel existing harus mengandung:  
  Harus di luar dari Nama kabel new
  Contoh: `DS-CKA-01`,`DS`,`DIS EXISTING` 

#### **D. OTB (Optical Termination Box)**
- Format: `OTB [NEW/BARU]`  
  Contoh:  
  ✅ `OTB 12 NEW`  
  ✅ `OTB BARU`  
- Langsung masuk ke designator **TC-SM-12**

#### **E. Closure**
- Nama harus mengandung:  
  ✅ `CL`, `CLOSURE`, atau `CL NEW`  
  Contoh: `CLOSURE NEW`

### **3. Troubleshooting**
- ❌ "Tidak Terdeteksi"? Periksa:  
  1. Penamaan sesuai format  
  2. Tipe geometry (Point/LineString) benar  
  3. Koordinat valid  
- 🔄 Jika error, export ulang dari Google Earth
""")

def manual_input_form():
    """Manual input form for BOQ generation"""
    initialize_session_state()
    
    with st.form("manual_boq_form"):
        st.subheader("Project Information")
        col1, col2 = st.columns([2, 1])
        with col1:
            st.session_state.boq_form_values['lop_name'] = st.text_input(
                "Nama LOP*",
                value=st.session_state.boq_form_values.get('lop_name', ""),
                help="Contoh: LOP_JAKARTA_123"
            )
        with col2:
            st.session_state.boq_form_values['sumber'] = st.radio(
                "Sumber*",
                ["ODC", "ODP"],
                index=0 if st.session_state.boq_form_values.get('sumber', "ODC") == "ODC" else 1,
                horizontal=True
            )

        st.subheader("Material Requirements")
        col1, col2 = st.columns(2)
        with col1:
            st.session_state.boq_form_values['kabel_12'] = st.number_input(
                "12 Core Cable (meter)*",
                min_value=0.0,
                value=st.session_state.boq_form_values.get('kabel_12', 0.0),
                step=1.0,
                format="%.1f"
            )
            st.session_state.boq_form_values['odp_8'] = st.number_input(
                "ODP 8 Port*",
                min_value=0,
                value=st.session_state.boq_form_values.get('odp_8', 0)
            )
            st.session_state.boq_form_values['otb_12'] = st.number_input(
                "OTB 12 (TC-SM-12)",
                min_value=0,
                value=st.session_state.boq_form_values.get('otb_12', 0)
            )
            st.session_state.boq_form_values['tiang_new'] = st.number_input(
                "Tiang Baru*",
                min_value=0,
                value=st.session_state.boq_form_values.get('tiang_new', 0)
            )
            
        with col2:
            st.session_state.boq_form_values['kabel_24'] = st.number_input(
                "24 Core Cable (meter)*",
                min_value=0.0,
                value=st.session_state.boq_form_values.get('kabel_24', 0.0),
                step=1.0,
                format="%.1f"
            )
            st.session_state.boq_form_values['odp_16'] = st.number_input(
                "ODP 16 Port*",
                min_value=0,
                value=st.session_state.boq_form_values.get('odp_16', 0)
            )
            st.session_state.boq_form_values['tiang_existing'] = st.number_input(
                "Tiang Eksisting*",
                min_value=0,
                value=st.session_state.boq_form_values.get('tiang_existing', 0)
            )
            st.session_state.boq_form_values['tikungan'] = st.number_input(
                "Tikungan*",
                min_value=0,
                value=st.session_state.boq_form_values.get('tikungan', 0)
            )

        st.subheader("Additional Inputs")
        col1, col2 = st.columns(2)
        with col1:
            st.session_state.boq_form_values['closure'] = st.number_input(
                "Closure (SC-OF-SM-24)",
                min_value=0,
                value=st.session_state.boq_form_values.get('closure', 0)
            )
        with col2:
            st.session_state.boq_form_values['izin'] = st.text_input(
                "Preliminary (isi nominal jika ada)",
                value=st.session_state.boq_form_values.get('izin', ""),
                help="Contoh: 500000"
            )

        st.subheader("Template File")
        st.session_state.boq_form_values['uploaded_file'] = st.file_uploader(
            "Unggah Template BOQ*",
            type=["xlsx"],
            help="Format Excel template BOQ"
        )

        submitted = st.form_submit_button("🚀 Generate BOQ Manual", use_container_width=True)
        
        if submitted:
            if not st.session_state.boq_form_values.get('uploaded_file'):
                st.error("Silakan unggah file template BOQ!")
                return
            if not st.session_state.boq_form_values.get('lop_name'):
                st.error("Silakan isi nama LOP!")
                return
            
            st.session_state.boq_state['active_tab'] = "manual"
            result = process_boq_template(
                st.session_state.boq_form_values['uploaded_file'],
                st.session_state.boq_form_values,
                st.session_state.boq_form_values['lop_name']
            )
            
            if result:
                st.session_state.boq_state.update({
                    'ready': True,
                    'excel_data': result['excel_data'],
                    'project_name': st.session_state.boq_form_values['lop_name'],
                    'updated_items': result['updated_items'],
                    'summary': result['summary']
                })
                st.success("✅ BOQ berhasil digenerate!")
                st.balloons()

def kml_input_form():
    """KML input form for BOQ generation"""
    initialize_session_state()
    
    with st.form("kml_boq_form"):
        st.subheader("Project Information")
        col1, col2 = st.columns([2, 1])
        with col1:
            st.session_state.boq_form_values['lop_name'] = st.text_input(
                "Nama LOP*",
                value=st.session_state.boq_form_values.get('lop_name', ""),
                key='kml_lop_name'
            )
        with col2:
            st.session_state.boq_form_values['sumber'] = st.radio(
                "Sumber*",
                ["ODC", "ODP"],
                index=0 if st.session_state.boq_form_values.get('sumber', "ODC") == "ODC" else 1,
                key='kml_sumber',
                horizontal=True
            )

        st.subheader("KML File Upload")
        display_kml_guide()
        
        st.session_state.boq_form_values['kml_file'] = st.file_uploader(
            "Unggah File KML*",
            type=["kml"],
            key='kml_uploader',
            help="File harus berisi: ODP NEW/BARU, Tiang, dan jalur kabel"
        )

        if st.session_state.boq_form_values.get('kml_file'):
            with st.spinner("Memproses KML..."):
                kml_values = parse_kml_file(st.session_state.boq_form_values['kml_file'])
                if kml_values:
                    st.success("✅ KML berhasil diproses!")
                    
                    st.session_state.boq_form_values.update({
                        'tiang_new': kml_values['tiang_new'],
                        'tiang_existing': kml_values['tiang_existing'],
                        'kabel_12': kml_values['kabel_12'],
                        'odp_8': kml_values['odp_8'],
                        'odp_16': kml_values['odp_16'],
                        'closure': kml_values['closure'],
                        'otb_12': kml_values['otb_12']
                    })

                    with st.expander("🔍 Hasil Deteksi KML"):
                        cols = st.columns(2)
                        with cols[0]:
                            st.metric("ODP 8 Port (NEW/BARU)", kml_values['odp_8'])
                            st.metric("Tiang Baru", kml_values['tiang_new'])
                            st.metric("Panjang Kabel (m)", f"{kml_values['kabel_12']:.2f}")
                        with cols[1]:
                            st.metric("ODP 16 Port (NEW/BARU)", kml_values['odp_16'])
                            st.metric("Tiang Existing", kml_values['tiang_existing'])
                            st.metric("OTB 12 (NEW/BARU)", kml_values['otb_12'])

        st.subheader("Additional Inputs")
        col1, col2 = st.columns(2)
        with col1:
            st.session_state.boq_form_values['tikungan'] = st.number_input(
                "Tikungan*",
                min_value=0,
                value=st.session_state.boq_form_values.get('tikungan', 0),
                key='kml_tikungan'
            )
        with col2:
            st.session_state.boq_form_values['izin'] = st.text_input(
                "Preliminary (isi nominal jika ada)",
                value=st.session_state.boq_form_values.get('izin', ""),
                key='kml_izin',
                help="Contoh: 500000"
            )

        st.subheader("Template File")
        st.session_state.boq_form_values['uploaded_file'] = st.file_uploader(
            "Unggah Template BOQ*",
            type=["xlsx"],
            key='kml_template'
        )

        submitted = st.form_submit_button("🚀 Generate BOQ dari KML", use_container_width=True)
        
        if submitted:
            if not st.session_state.boq_form_values.get('uploaded_file'):
                st.error("Silakan unggah file template BOQ!")
                return
            if not st.session_state.boq_form_values.get('lop_name'):
                st.error("Silakan isi nama LOP!")
                return
            if not st.session_state.boq_form_values.get('kml_file'):
                st.error("Silakan unggah file KML!")
                return
            
            st.session_state.boq_state['active_tab'] = "kml"
            result = process_boq_template(
                st.session_state.boq_form_values['uploaded_file'],
                st.session_state.boq_form_values,
                st.session_state.boq_form_values['lop_name']
            )
            
            if result:
                st.session_state.boq_state.update({
                    'ready': True,
                    'excel_data': result['excel_data'],
                    'project_name': st.session_state.boq_form_values['lop_name'],
                    'updated_items': result['updated_items'],
                    'summary': result['summary']
                })
                st.success("✅ BOQ berhasil digenerate!")
                st.balloons()

def display_results():
    """Display the BOQ generation results"""
    st.divider()
    st.subheader("📊 Hasil BOQ")
    
    summary = st.session_state.boq_state['summary']
    cols = st.columns(4)
    with cols[0]:
        st.metric("Total ODP", summary['total_odp'])
        st.metric("Total Port", summary['total_ports'])
    with cols[1]:
        st.metric("Material", f"Rp {summary['material']:,.0f}")
    with cols[2]:
        st.metric("Jasa", f"Rp {summary['jasa']:,.0f}")
    with cols[3]:
        st.metric("Total Biaya", f"Rp {summary['total']:,.0f}")
        st.metric("CPP", f"Rp {summary['cpp']:,.0f}")
    
    st.subheader("📋 Item yang Diupdate")
    df_items = pd.DataFrame(st.session_state.boq_state['updated_items'])
    st.dataframe(df_items, hide_index=True, use_container_width=True)
    
    st.subheader("📥 Download")
    col1, col2 = st.columns([1, 2])
    with col1:
        st.download_button(
            label="⬇️ Download BOQ",
            data=st.session_state.boq_state['excel_data'],
            file_name=f"BOQ-{st.session_state.boq_state['project_name']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col2:
        if st.button("🔄 Buat BOQ Baru", use_container_width=True):
            reset_boq_application()
            st.rerun()

def main():
    """Main application function"""
    st.set_page_config(
        page_title="BOQ Generator",
        page_icon="📊",
        layout="wide"
    )
    
    # Custom CSS
    st.markdown("""
    <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 10px;
        }
        .stTabs [data-baseweb="tab"] {
            padding: 8px 16px;
            border-radius: 4px 4px 0 0;
        }
        .stAlert {padding: 20px;}
    </style>
    """, unsafe_allow_html=True)
    
    initialize_session_state()
    
    st.title("📊 BOQ Generator")
    st.markdown("Aplikasi untuk menghasilkan Bill of Quantities (BOQ) proyek fiber optik")
    
    tab1, tab2 = st.tabs(["📝 Manual Input", "🗺️ BOQ dari KML"])
    
    with tab1:
        manual_input_form()
    
    with tab2:
        kml_input_form()
    
    if st.session_state.boq_state.get('ready', False):
        display_results()

if __name__ == "__main__":
    main()
